import sys
import logging
import requests
from requests.exceptions import RequestException, ConnectionError, HTTPError, Timeout
from bs4 import BeautifulSoup
from pathlib import Path
import openpyxl

# 一次元配列を二次元配列に変換(colsに要素数を設定)
def convert_1d_to_2d(l, cols):
    return [l[i:i + cols] for i in range(0, len(l), cols)]

# log出力の設定
folder = Path('log')
folder.mkdir(exist_ok=True)
logging.basicConfig(filename=folder.joinpath('log.txt'), level=logging.INFO)
logger = logging.getLogger(__name__)

# 起動時引数（検索サイトURL）の取得
if len(sys.argv) > 1:
    url = sys.argv[1]
else:
    logger.error('Site specification argument error')
    sys.exit()
 # サイト情報取得・解析
try:
    soup = BeautifulSoup(requests.get(url).content, 'html.parser')

    title_tags = soup.select('h4[id^=section]')
    table_tags = soup.select('div[id^=tab01_]')
except ConnectionError as e:        # インターネット接続エラー
    logger.exception(f'ConnectionError:{e}')
    sys.exit()
except HTTPError as e:              # HTTOステータスエラー
    logger.exception(f'HTTPError:{e}')
    sys.exit()
except Timeout as e:                # リクエストタイムアウト
    logger.exception(f'Timeout:{e}')
    sys.exit()
except RequestException as e:       # その他例外発生
    logger.exception(f'Error:{e}')
    sys.exit()

logger.info('Request successfully completed')

# excelブック/シートの準備
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'info_slot'
ws.cell(row=1, column=1, value='設置機種')

tbl = dict()
counter = 1

for num in range(len(title_tags)):
    if num != len(title_tags) - 1:
        tbl['title'] = title_tags[num].text
        header_tags = table_tags[num].select('th')
        tbl['header'] = [header_tag.text for header_tag in header_tags]
        detail_tags = table_tags[num].select('td')
        tbl['detail'] = convert_1d_to_2d([detail_tag.text for detail_tag in detail_tags], len(tbl['header']))

        if num == 0:
            ws.cell(row=counter, column=2, value='台番号')
            ws.cell(row=counter, column=3, value='G数')
            ws.cell(row=counter, column=4, value='BB')
            ws.cell(row=counter, column=5, value='RB')
            ws.cell(row=counter, column=6, value='差枚')
            counter += 1

        ws.cell(row=counter, column=1, value=tbl['title'])

        for row in tbl['detail']:
            if row[0] != '平均':
                ws.cell(row=counter, column=2, value=row[0])
                ws.cell(row=counter, column=3, value=row[1])
                ws.cell(row=counter, column=4, value=row[3])
                ws.cell(row=counter, column=5, value=row[4])
                ws.cell(row=counter, column=6, value=row[2])
                counter += 1

        counter += 1
    else:
        tbl['title'] = title_tags[num].text
        header_tags = table_tags[num].select('th')
        tbl['header'] = [header_tag.text for header_tag in header_tags]
        detail_tags = table_tags[num].select('td')
        tbl['detail'] = convert_1d_to_2d([detail_tag.text for detail_tag in detail_tags], len(tbl['header']))

        for row in tbl['detail']:
            ws.cell(row=counter, column=1, value=row[0])
            ws.cell(row=counter, column=2, value=row[1])
            ws.cell(row=counter, column=3, value=row[2])
            ws.cell(row=counter, column=4, value=row[4])
            ws.cell(row=counter, column=5, value=row[5])
            ws.cell(row=counter, column=6, value=row[3])
            counter += 1

    # 列幅調整
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1) * 2
        ws.column_dimensions[column].width = adjusted_width

# 出力フォルダーを作成
folder = Path('output')
folder.mkdir(exist_ok=True)

# excelファイルに出力
# excel_path = folder.joinpath('info_slot.xlsx')
excel_path = folder.joinpath(url[20:30]+'.xlsx')
wb.save(excel_path)
wb.close()