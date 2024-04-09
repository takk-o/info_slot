import requests
from requests.exceptions import RequestException, ConnectionError, HTTPError, Timeout
from bs4 import BeautifulSoup
from pathlib import Path
import openpyxl
import tkinter as tk
from pathlib import Path
import os
import sys

# 一次元配列を二次元配列に変換(colsに要素数を設定)
def convert_1d_to_2d(l, cols):
    return [l[i:i + cols] for i in range(0, len(l), cols)]

def send_message_label(msg):
    label3['text'] = label3['text'] + msg + '\n'
    label3.update()

def info_slot_main(url):

    send_message_label(f'{url} へのアクセス開始')

    # サイト情報取得・解析
    try:
        soup = BeautifulSoup(requests.get(url).content, 'html.parser')

        title_tags = soup.select('h4[id^=section]')
        table_tags = soup.select('div[id^=tab01_]')
    except ConnectionError as e:        # インターネット接続エラー
        send_message_label(f'ConnectionError:{e}')
        return
    except HTTPError as e:              # HTTOステータスエラー
        send_message_label(f'HTTPError:{e}')
        return
    except Timeout as e:                # リクエストタイムアウト
        send_message_label(f'Timeout:{e}')
        return
    except RequestException as e:       # その他例外発生
        send_message_label(f'Error:{e}')
        return
    
    send_message_label('Excel編集開始')

    # excelブック/シートの準備
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'info_slot'
    ws.cell(row=1, column=1, value='設置機種')

    tbl = dict()
    counter = 1

    for num in range(len(title_tags)):
        if num != len(title_tags) - 1:
            tbl['title'] = title_tags[num].text
            header_tags = table_tags[num].select('th')
            tbl['header'] = [header_tag.text for header_tag in header_tags]
            detail_tags = table_tags[num].select('td')
            tbl['detail'] = convert_1d_to_2d([detail_tag.text for detail_tag in detail_tags], len(tbl['header']))

            if num == 0:
                ws.cell(row=counter, column=2, value='台番号')
                ws.cell(row=counter, column=3, value='G数')
                ws.cell(row=counter, column=4, value='BB')
                ws.cell(row=counter, column=5, value='RB')
                ws.cell(row=counter, column=6, value='差枚')
                counter += 1

            ws.cell(row=counter, column=1, value=tbl['title'])

            for row in tbl['detail']:
                if row[0] != '平均':
                    ws.cell(row=counter, column=2, value=row[0])
                    ws.cell(row=counter, column=3, value=row[1])
                    ws.cell(row=counter, column=4, value=row[3])
                    ws.cell(row=counter, column=5, value=row[4])
                    ws.cell(row=counter, column=6, value=row[2])
                    counter += 1

            counter += 1
        else:
            tbl['title'] = title_tags[num].text
            header_tags = table_tags[num].select('th')
            tbl['header'] = [header_tag.text for header_tag in header_tags]
            detail_tags = table_tags[num].select('td')
            tbl['detail'] = convert_1d_to_2d([detail_tag.text for detail_tag in detail_tags], len(tbl['header']))

            for row in tbl['detail']:
                ws.cell(row=counter, column=1, value=row[0])
                ws.cell(row=counter, column=2, value=row[1])
                ws.cell(row=counter, column=3, value=row[2])
                ws.cell(row=counter, column=4, value=row[4])
                ws.cell(row=counter, column=5, value=row[5])
                ws.cell(row=counter, column=6, value=row[3])
                counter += 1

        # 列幅調整
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1) * 2
            ws.column_dimensions[column].width = adjusted_width

    send_message_label('Excel編集終了')

    # 出力フォルダーを作成
    folder = Path('output')
    folder.mkdir(exist_ok=True)

    # excelファイルに出力
    fname = url.split('/')[3][:-5]
    excel_path = folder.joinpath(fname +'.xlsx')
    wb.save(excel_path)
    wb.close()

    send_message_label('Excel出力終了')
    entry_box.delete(0, tk.END)
    entry_box.update()

# os.chdir(os.environ['HOME'])
# os.chdir('/Users/hisui')
os.chdir(Path(sys.argv[0]).resolve().parents[3])

root = tk.Tk()
root.title('info_slot')
root.resizable(False, True)
root.rowconfigure(1, weight=1)

label1 = tk.Label(root, text='URL')
label1.grid(row=0, column=0, columnspan=1)

entry_box = tk.Entry(root, width=100, fg='black', bg='white')
entry_box.grid(row=0, column=1, columnspan=1, sticky=tk.NE+tk.SW)

button1 = tk.Button(root, text='実行', bg='#ff7f50', command=lambda:info_slot_main(entry_box.get()))
button1.grid(row=0, column=2, columnspan=1)

label2 = tk.Label(root, text='進行状況')
label2.grid(row=1, column=0, columnspan=1)

label3 = tk.Label(root, text=f'出力フォルダ：{Path.cwd()}\n', height=5, justify='left', fg='black', bg='lavender', anchor=tk.NW)
label3.grid(row=1, column=1, columnspan=2, sticky=tk.NE+tk.SW)

button2 = tk.Button(root, text='終了', bg='#ff7f50', command=root.destroy)
button2.grid(row=2, column=2, columnspan=1)

root.mainloop()